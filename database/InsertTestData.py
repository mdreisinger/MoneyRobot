"""
Module to insert test data into the moneyrobot database.
"""
import json
import os
import pathlib
import sys

from aws_secretsmanager_caching import SecretCache, SecretCacheConfig
import botocore
import botocore.session
import click
from openpyxl import load_workbook
import pymysql


@click.command()
def insert_test_data():
    """
    Function to insert test data into the moneyrobot database.
    """
    rds_host = os.getenv("RDS_HOST")
    username = os.getenv("USERNAME")
    password = os.getenv("PASSWORD")
    db_name = os.getenv("DB_NAME")

    if rds_host and username and password and db_name:
        print("Using environment variables for database connection info!")
    else:
        print("Creating boto client to secretsmanager to get database connection info!")
        client = botocore.session.get_session().create_client('secretsmanager', region_name='us-west-2')
        cache_config = SecretCacheConfig()
        cache = SecretCache( config = cache_config, client = client)
        secret = json.loads(cache.get_secret_string('moneyrobot-dev-secret'))
        rds_host  = secret["host"]
        username = secret["username"]
        password = secret["password"]
        db_name = secret["dbname"]

    print(f"Connecting to {db_name} on {rds_host}!")
    try:
        conn = pymysql.connect(host=rds_host,
                               user=username,
                               passwd=password,
                               db=db_name,
                               connect_timeout=10,
                               cursorclass=pymysql.cursors.DictCursor)
    except pymysql.MySQLError as error:
        print("ERROR: Unexpected error: Could not connect to MySQL instance.")
        print(error)
        sys.exit()
    print("SUCCESS: Connection to RDS MySQL instance succeeded")

    cur_dir = pathlib.Path(__file__).parent.resolve()
    book = load_workbook(f"{cur_dir}/sample_data.xlsx")
    sheet = book.active

    with conn.cursor() as cur:
        for row in range(1, sheet.max_row + 1):
            expense_date = str(sheet.cell(row, 1).value).replace("'","")
            payee = str(sheet.cell(row, 2).value).replace("'","")
            items = str(sheet.cell(row, 3).value).replace("'","")
            expense = str(sheet.cell(row, 4).value).replace("'","")
            sql = "INSERT INTO expenses (expense_date, payee, items, expense) VALUES "
            sql += f"('{expense_date}', "
            sql += f"'{payee}', "
            sql += f"'{items}', "
            sql += f"'{expense}')"
            cur.execute(sql)
            print(sql)
    conn.commit()
