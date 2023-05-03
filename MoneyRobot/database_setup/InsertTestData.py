"""
Module to insert test data into the moneyrobot database.
"""
import pathlib

import click
from openpyxl import load_workbook

from .DatabaseConnection import get_connection


@click.command()
def insert_test_data():
    """
    Function to insert test data into the moneyrobot database.
    """

    conn = get_connection()
    
    cur_dir = pathlib.Path(__file__).parent.resolve()
    book = load_workbook(f"{cur_dir}/sample_transactions.xlsx")
    sheet = book.active

    with conn.cursor() as cur:
        for row in range(1, sheet.max_row + 1):
            transaction_date = str(sheet.cell(row, 1).value).replace("'","")
            transactor = str(sheet.cell(row, 2).value).replace("'","")
            transaction_category = str(sheet.cell(row, 3).value).replace("'","")
            items = str(sheet.cell(row, 4).value).replace("'","")
            transaction_amount = str(sheet.cell(row, 5).value).replace("'","")
            sql = "INSERT INTO transactions (transaction_date, transactor, transaction_category, items, transaction_amount) VALUES "
            sql += f"('{transaction_date}', "
            sql += f"'{transactor}', "
            sql += f"'{transaction_category}', "
            sql += f"'{items}', "
            sql += f"'{transaction_amount}')"
            cur.execute(sql)
            print(sql)
    conn.commit()
